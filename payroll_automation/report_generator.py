import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import contextlib

@contextlib.contextmanager
def managed_workbook(filename):
    wb = Workbook()
    try:
        yield wb
    finally:
        wb.save(filename)
        wb.close()

def generate_excel_report(salary_info, processed_timesheet, vacation_info, schedule_changes, schedule_alerts, output_file):
    with managed_workbook(output_file) as wb:
        # 1. Summary Salary Report
        ws = wb.active
        ws.title = "Summary Salary Report"
        
        # Add headers
        headers = ["Employee Number", "Employee Name", "Department", "Total Hours", "Regular Pay", "Overtime Pay", "Holiday Pay", "Bonus", "Total Compensation"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row, (_, employee) in enumerate(salary_info.iterrows(), start=2):
            employee_timesheet = combined_timesheet[combined_timesheet['Employee Number'] == employee['Employee Number']]
            department = employee_timesheet['Department'].iloc[0] if not employee_timesheet.empty else 'N/A'
            
            ws.cell(row=row, column=1, value=employee['Employee Number'])
            ws.cell(row=row, column=2, value=employee['Employee Name'])
            ws.cell(row=row, column=3, value=department)
            ws.cell(row=row, column=4, value=employee['Total Hours'])
            ws.cell(row=row, column=5, value=employee['Salary'] - employee['OT Pay Rate (加班时薪）'] * max(0, employee['Total Hours'] - employee['Bi-weekly 加班费触发小时（有holiday）']))
            ws.cell(row=row, column=6, value=employee['OT Pay Rate (加班时薪）'] * max(0, employee['Total Hours'] - employee['Bi-weekly 加班费触发小时（有holiday）']))
            ws.cell(row=row, column=7, value=employee['Holiday Pay 08-05'] if 'Holiday Pay 08-05' in employee else 0)
            ws.cell(row=row, column=8, value=employee['Bonus'])
            ws.cell(row=row, column=9, value=employee['Total Compensation'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # # 2. Individual Payroll Reports
        for _, employee in salary_info.iterrows():
            sheet_name = f"{employee['Employee Number']} - {employee['Employee Name']}"[:31]  # Excel has a 31 character limit for sheet names
            ws = wb.create_sheet(title=sheet_name)
            
            employee_timesheet = processed_timesheet[processed_timesheet['Employee Number'] == employee['Employee Number']].copy()
            department = employee_timesheet['Department'].iloc[0] if not employee_timesheet.empty else 'N/A'
            
            # Employee Information
            ws['A1'] = "Employee Payroll Report"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A3'] = f"Employee Number: {employee['Employee Number']}"
            ws['A4'] = f"Name: {employee['Employee Name']}"
            ws['A5'] = f"Department: {department}"
            
            # Payroll Summary
            ws['A7'] = "Payroll Summary"
            ws['A7'].font = Font(bold=True)
            ws['A8'] = "Total Hours:"
            ws['B8'] = employee['Total Hours']
            ws['A9'] = "Regular Pay Rate:"
            ws['B9'] = employee['REG Pay Rate (正常时薪)']
            ws['A10'] = "Regular Pay:"
            regular_pay = employee['Salary'] - employee.get('OT Pay Rate (加班时薪）', 0) * max(0, employee['Total Hours'] - employee.get('Bi-weekly 加班费触发小时（有holiday）', 0))
            ws['B10'] = regular_pay
            ws['A11'] = "Overtime Pay Rate:"
            ws['B11'] = employee.get('OT Pay Rate (加班时薪）', 0)
            ws['A12'] = "Overtime Pay:"
            overtime_pay = employee.get('OT Pay Rate (加班时薪）', 0) * max(0, employee['Total Hours'] - employee.get('Bi-weekly 加班费触发小时（有holiday）', 0))
            ws['B12'] = overtime_pay
            ws['A13'] = "Holiday Pay:"
            ws['B13'] = employee.get('Holiday Pay 08-05', 0)
            ws['A14'] = "Bonus:"
            ws['B14'] = employee.get('Bonus', 0)
            ws['A15'] = "Total Compensation:"
            ws['B15'] = employee['Total Compensation']
            
            # Additional Information
            ws['A17'] = "Additional Information"
            ws['A17'].font = Font(bold=True)
            ws['A18'] = "Bi-weekly Overtime Threshold (with holiday):"
            ws['B18'] = employee.get('Bi-weekly 加班费触发小时（有holiday）', 'N/A')
            ws['A19'] = "Bi-weekly Overtime Threshold (without holiday):"
            ws['B19'] = employee.get('Bi-weekly 加班费触发小时（没有holiday）', 'N/A')
            ws['A20'] = "Annual Or Hourly:"
            ws['B20'] = employee.get('Annual Or Hourly', 'N/A')
            ws['A21'] = "Follow 打卡时间:"
            ws['B21'] = employee.get('Follow 打卡时间', 'N/A')
            
            # Select and reorder columns for the timesheet
            timesheet_columns = ['Employee Number', 'First name', 'Last name', 'Job', 'Department', 
                                 'Start Datetime', 'End Datetime', 'Adjusted Start Datetime', 
                                 'Adjusted End Datetime', 'Working Hours']
            
            employee_timesheet_filtered = employee_timesheet[timesheet_columns]
            
            # Add timesheet data
            for r in dataframe_to_rows(employee_timesheet_filtered, index=False, header=True):
                ws.append(r)

            # Auto-adjust column widths for timesheet
            for col_num, column_title in enumerate(timesheet_columns, start=1):
                column_letter = get_column_letter(col_num)
                max_length = len(str(column_title))
                for cell in ws[column_letter][24:]:  # Start from row 24 (after the header)
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width

        
        # 3. Vacation Report
        ws = wb.create_sheet(title="Vacation Report")
        
        # Add headers
        headers = ["Employee Number", "Employee Name", "Department", "Vacation Days"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Add data
        for row, (_, vacation) in enumerate(vacation_info.iterrows(), start=2):
            ws.cell(row=row, column=1, value=vacation['Employee Number'])
            ws.cell(row=row, column=2, value=f"{vacation['First Name']} {vacation['Last Name']}")
            ws.cell(row=row, column=3, value=vacation['Department'])
            ws.cell(row=row, column=4, value=vacation['Vacation Days'])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

         # 4. Schedule Changes Report
        if isinstance(schedule_changes, list) and schedule_changes:
            ws = wb.create_sheet(title="Schedule Changes")
            
            # Add headers
            headers = ["Employee Number", "Full Name", "Original Start", "New Start", "Time Difference (hours)", "Reason"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Add data
            for row, change in enumerate(schedule_changes, start=2):
                if isinstance(change, dict):
                    for col, key in enumerate(headers, start=1):
                        ws.cell(row=row, column=col, value=change.get(key, ''))
                else:
                    print(f"Warning: Unexpected type in schedule_changes: {type(change)}")
        else:
            print("No schedule changes to report or invalid data type.")

        # 5. Schedule Alerts Report
        if isinstance(schedule_alerts, list) and schedule_alerts:
            ws = wb.create_sheet(title="Schedule Alerts")
            
            # Add headers
            headers = ["Employee Number", "Full Name", "Timesheet Start", "Schedule Start", "Time Difference (hours)", "Reason"]
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Add data
            for row, alert in enumerate(schedule_alerts, start=2):
                if isinstance(alert, dict):
                    for col, key in enumerate(headers, start=1):
                        ws.cell(row=row, column=col, value=alert.get(key, ''))
                else:
                    print(f"Warning: Unexpected type in schedule_alerts: {type(alert)}")
        else:
            print("No schedule alerts to report or invalid data type.")

        # Auto-adjust column widths for new sheets
        for sheet_name in ["Schedule Changes", "Schedule Alerts"]:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(cell.value)
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column_letter].width = adjusted_width

    print(f"Excel report generated: {output_file}")

    print(f"Excel report generated: {output_file}")
